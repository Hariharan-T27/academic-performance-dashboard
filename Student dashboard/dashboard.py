import streamlit as st
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from streamlit_option_menu import option_menu
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# -------------------------------
# Page Config
# -------------------------------
st.set_page_config(page_title="Student Performance Dashboard", layout="wide")

# -------------------------------
# Load Dataset
# -------------------------------
df = pd.read_csv("student_performance.csv")

# Ensure StudentID column exists
if "StudentID" not in df.columns:
    df.insert(0, "StudentID", [f"S{str(i+1).zfill(3)}" for i in range(len(df))])
else:
    cols = ['StudentID'] + [col for col in df.columns if col != 'StudentID']
    df = df[cols]

# Function to generate next Student ID
def generate_student_id(df):
    if df.empty:
        return "S001"
    last_id = df["StudentID"].iloc[-1]
    last_num = int(last_id[1:])
    return f"S{str(last_num + 1).zfill(3)}"

# -------------------------------
# Custom CSS
# -------------------------------
st.markdown("""
    <style>
    .main { background-color: #F9F9F9; }
    .metric-card {
        background-color: #FFFFFF;
        color: #1A1A1A;
        padding: 20px;
        border-radius: 16px;
        box-shadow: 0 8px 28px rgba(0,0,0,0.12), 0 2px 6px rgba(0,0,0,0.6);
        text-align: center;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .metric-card:hover{
        transform: translateY(-4px);
        box-shadow: 0 8px 28px rgba(0, 0, 0, 0.12),
                    0 4px 12px rgba(0, 0, 0, 0.08);
    }                
    .metric-card h2 { font-size: 28px; font-weight: bold; }
    .metric-card h3 { font-size: 18px; }
    </style>
""", unsafe_allow_html=True)

# -------------------------------
# Header
# -------------------------------
st.markdown("""
    <div style="background-color:#065F46;padding:15px;border-radius:10px;margin-bottom:20px">
    <h1 style="color:white;text-align:center;">📊 Student Performance Analytics Dashboard</h1>
    </div>
""", unsafe_allow_html=True)

# -------------------------------
# Sidebar with Risk Thresholds
# -------------------------------
with st.sidebar:
    choice = option_menu("Dashboard Sections",
                         ["Overview", "Correlation Analysis", "Student Insights"],
                         icons=["house", "bar-chart", "people"],
                         menu_icon="cast", default_index=0)

    st.write("### 📌 Risk Thresholds")
    high_risk_cutoff = st.slider("Marks cutoff for High Risk", 0, 100, 40, key="high_risk_slider")
    medium_risk_cutoff = st.slider("Marks cutoff for Medium Risk", 0, 100, 60, key="medium_risk_slider")
    attendance_cutoff = st.slider("Attendance cutoff (%)", 0, 100, 50, key="attendance_slider")

# Initialize session state
if "pending_student" not in st.session_state:
    st.session_state.pending_student = None

# Ensure Risk column exists always
df["Risk"] = df.apply(
    lambda x: "High Risk" if (x["Marks"] < high_risk_cutoff or x["Attendance"] < attendance_cutoff)
    else "Medium Risk" if x["Marks"] < medium_risk_cutoff
    else "Low Risk", axis=1
)

# -------------------------------
# Overview Section
# -------------------------------
if choice == "Overview":
    st.subheader("📌 Overview Metrics")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"<div class='metric-card'><h3>Average Marks</h3><h2>{df['Marks'].mean():.2f}</h2></div>", unsafe_allow_html=True)
    with col2:
        st.markdown(f"<div class='metric-card'><h3>Average Attendance</h3><h2>{df['Attendance'].mean():.1f}%</h2></div>", unsafe_allow_html=True)
    with col3:
        st.markdown(f"<div class='metric-card'><h3>Average Logins</h3><h2>{df['Logins'].mean():.0f}</h2></div>", unsafe_allow_html=True)
    with col4:
        st.markdown(f"<div class='metric-card'><h3>Total Students</h3><h2>{len(df)}</h2></div>", unsafe_allow_html=True)

    st.write("---")
    st.subheader("📝 New Student Admission Form")

    with st.form("admission_form", clear_on_submit=False):
        name = st.text_input("Full Name*", placeholder="Enter student's full name")
        marks = st.number_input("Marks", min_value=0, max_value=100, step=1)
        attendance = st.number_input("Attendance (%)", min_value=0, max_value=100, step=1)
        logins = st.number_input("Logins", min_value=0, step=1)
        submitted = st.form_submit_button("Preview Student")

        if submitted:
            if not name.strip():
                st.error("⚠️ Name is required!")
            else:
                duplicate = df[
                    (df['Name'].str.lower() == name.strip().lower()) &
                    (df['Marks'] == marks) &
                    (df['Attendance'] == attendance) &
                    (df['Logins'] == logins)
                ]
                if not duplicate.empty:
                    st.warning("⚠️ This student entry already exists!")
                else:
                    new_id = generate_student_id(df)
                    risk = "High Risk" if (marks < high_risk_cutoff or attendance < attendance_cutoff) \
                           else "Medium Risk" if marks < medium_risk_cutoff else "Low Risk"
                    st.session_state.pending_student = {
                        "StudentID": new_id,
                        "Name": name.strip(),
                        "Marks": marks,
                        "Attendance": attendance,
                        "Logins": logins,
                        "Risk": risk
                    }

# -------------------------------
# Confirmation Card
# -------------------------------
if st.session_state.pending_student:
    student = st.session_state.pending_student
    color_map = {"High Risk": "#F1948A", "Medium Risk": "#F8C471", "Low Risk": "#82E0AA"}
    st.markdown(f"""
        <div style="border-radius:10px; padding:15px; background-color:{color_map[student['Risk']]};
                    text-align:center; margin-top:20px;">
            <h3 style="color:black;">Confirm New Student Entry</h3>
            <b>Student ID:</b> {student['StudentID']}<br>
            <b>Name:</b> {student['Name']}<br>
            <b>Marks:</b> {student['Marks']}<br>
            <b>Attendance:</b> {student['Attendance']}%<br>
            <b>Logins:</b> {student['Logins']}<br>
            <b>Risk:</b> {student['Risk']}
        </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        if st.button("✅ Confirm Add", key="confirm_add"):
            df.loc[len(df)] = student
            df.to_csv("student_performance.csv", index=False)
            st.success(f"✅ Student '{student['Name']}' added with ID {student['StudentID']} and saved permanently!")
            st.session_state.pending_student = None
    with col2:
        if st.button("❌ Cancel", key="cancel_add"):
            st.info("Operation cancelled. Student not added.")
            st.session_state.pending_student = None

    st.write("---")
    st.subheader("📂 Dataset Preview")
    st.dataframe(df.tail(10))

    st.write("---")
    st.subheader("📥 Export Updated Dataset")

    # CSV Export
    csv = df.to_csv(index=False).encode('utf-8')
    st.download_button("⬇️ Download CSV", csv, "updated_student_performance.csv", "text/csv")

    # Excel Export
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="StudentData")
        summary_data = {
            "Metric": ["Total Students", "Average Marks", "Average Attendance", "Average Logins"],
            "Value": [
                len(df),
                round(df["Marks"].mean(), 2),
                round(df["Attendance"].mean(), 2),
                round(df["Logins"].mean(), 2),
            ],
        }
        pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name="Summary", startrow=1)
        risk_counts = df["Risk"].value_counts().reset_index()
        risk_counts.columns = ["Risk Category", "Number of Students"]
        risk_counts.to_excel(writer, index=False, sheet_name="Summary", startrow=7)

        workbook = writer.book
        data_ws = writer.sheets["StudentData"]
        data_ws.freeze_panes = "A2"

        # Color Risk column
        risk_col_idx = list(df.columns).index("Risk") + 1
        color_map_excel = {"High Risk": "FF9999", "Medium Risk": "FFD580", "Low Risk": "90EE90"}
        for row in range(2, len(df) + 2):
            risk_val = data_ws.cell(row=row, column=risk_col_idx).value
            fill_color = color_map_excel.get(risk_val, "FFFFFF")
            data_ws.cell(row=row, column=risk_col_idx).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Auto-adjust column widths
        for col in data_ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            data_ws.column_dimensions[col_letter].width = max_length + 2

    st.download_button("⬇️ Download Excel", excel_buffer.getvalue(),
                       "updated_student_performance.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



# -------------------------------
# Correlation Analysis
# -------------------------------
elif choice == "Correlation Analysis":
    col1, col2, col3 = st.columns(3)
    with col1:
        st.subheader("🔗 Correlation Heatmap")
        fig, ax = plt.subplots(figsize=(6, 4))
        sns.heatmap(df[['Marks', 'Attendance', 'Logins']].corr(), annot=True, cmap="Blues", ax=ax)
        st.pyplot(fig)
    with col2:
        st.subheader("📍 Attendance vs Marks Scatterplot")
        fig, ax = plt.subplots(figsize=(6, 4))
        sns.scatterplot(x='Attendance', y='Marks', data=df, hue='Marks', palette="viridis", ax=ax)
        st.pyplot(fig)
    with col3:
        st.subheader("📊 Absentee Impact Analysis")
        df['Attendance_Group'] = pd.cut(df['Attendance'], bins=[0, 50, 60, 70, 80, 90, 100],
                                        labels=["<50%", "50-60%", "60-70%", "70-80%", "80-90%", "90-100%"])
        impact = df.groupby('Attendance_Group')['Marks'].mean().reset_index()
        fig, ax = plt.subplots(figsize=(7, 4))
        sns.barplot(x='Attendance_Group', y='Marks', data=impact, palette="coolwarm", ax=ax)
        ax.set_title("Average Marks by Attendance Group")
        st.pyplot(fig)



# -------------------------------
# Student Insights
# -------------------------------
elif choice == "Student Insights":
    st.subheader("🔍 Search Student Records")
    search_query = st.text_input("Enter Student ID or Name:", placeholder="e.g., S005 or John Doe")

    if search_query.strip():
        filtered_df = df[
            df['StudentID'].str.contains(search_query.strip(), case=False, na=False) |
            df['Name'].str.contains(search_query.strip(), case=False, na=False)
        ]
        if not filtered_df.empty:
            st.success(f"✅ Found {len(filtered_df)} matching student(s).")
            st.dataframe(filtered_df[['StudentID','Name','Marks','Attendance','Logins','Risk']])
        else:
            st.warning("⚠️ No matching student found.")
        st.stop()
    else:
        st.subheader("🏆 Top Performing Students")
        top_students = df[df['Marks'] > 90][['StudentID','Name','Marks','Attendance','Logins','Risk']]
        st.dataframe(top_students)

        st.subheader("⚖️ Average Performing Students")
        avg_students = df[(df['Marks'] >= 40) & (df['Marks'] <= 90)][['StudentID','Name','Marks','Attendance','Logins','Risk']]
        st.dataframe(avg_students)

        st.subheader("⚠️ Struggling Students")
        low_students = df[df['Marks'] < 40][['StudentID','Name','Marks','Attendance','Logins','Risk']]
        st.dataframe(low_students)

        st.subheader("🚨 Risk Categorization")
        risk_count = df['Risk'].value_counts()
        col1, col2 = st.columns(2)
        with col1:
            st.bar_chart(risk_count)
        with col2:
            st.write("Risk Distribution", risk_count)

    # -------------------------------
    # Delete Student Section
    # -------------------------------
    import time

    st.write("---")
    st.subheader("🗑️ Delete Student Record")

    # Initialize state for delete mode and confirmation
    if "delete_mode" not in st.session_state:
        st.session_state.delete_mode = False
    if "delete_candidate" not in st.session_state:
        st.session_state.delete_candidate = None

    # Step 1: Show Delete button
    if not st.session_state.delete_mode:
        if st.button("🗑️ Delete Student"):
            st.session_state.delete_mode = True
            st.rerun()

    # Step 2: Show search bar for deletion
    if st.session_state.delete_mode:
        search_query = st.text_input("Enter Student ID or Name to delete:", placeholder="e.g., S005 or John Doe")

        if search_query.strip():
            # Find matching student(s)
            filtered_df = df[
                df['StudentID'].str.contains(search_query.strip(), case=False, na=False) |
                df['Name'].str.contains(search_query.strip(), case=False, na=False)
                ]

            if not filtered_df.empty:
                st.success(f"✅ Found {len(filtered_df)} matching student(s). Select carefully.")
                st.dataframe(filtered_df[['StudentID', 'Name', 'Marks', 'Attendance', 'Logins', 'Risk']])

                # Ask faculty to type the exact ID of the student they want to delete
                confirm_id = st.text_input("Confirm the Student ID you want to delete:")

                if confirm_id.strip():
                    candidate = df[df['StudentID'].str.upper() == confirm_id.strip().upper()]
                    if not candidate.empty:
                        student_name = candidate['Name'].values[0]
                        student_id = candidate['StudentID'].values[0]
                        st.session_state.delete_candidate = student_id

                        st.warning(f"⚠️ Are you sure you want to permanently delete {student_name} (ID: {student_id})?")
                        col1, col2 = st.columns(2)
                        with col1:
                            if st.button("✅ Yes, Delete Permanently"):
                                df = df[df['StudentID'].str.upper() != student_id]
                                df.to_csv("student_performance.csv", index=False)

                                # Create placeholder for popup
                                popup_placeholder = st.empty()
                                popup_placeholder.markdown(f"""
                                    <div style="background-color:#d4edda; border:1px solid #155724;
                                                border-radius:10px; padding:15px; margin-top:20px;">
                                        <h4 style="color:#155724;">✅ Deletion Successful</h4>
                                        <p style="color:#155724;">
                                            Student <b>{student_name}</b> with ID <b>{student_id}</b> 
                                            has been permanently deleted.
                                        </p>
                                        <p style="color:#6c757d; font-size:13px;">This will auto-close in 5 seconds...</p>
                                    </div>
                                """, unsafe_allow_html=True)

                                # Wait 5 seconds then clear popup
                                time.sleep(5)
                                popup_placeholder.empty()

                                st.session_state.delete_mode = False
                                st.session_state.delete_candidate = None
                                st.rerun()
                        with col2:
                            if st.button("❌ Cancel Deletion"):
                                st.info("Deletion cancelled.")
                                st.session_state.delete_mode = False
                                st.session_state.delete_candidate = None
                                st.rerun()

                    else:
                        st.error("⚠️ No matching Student ID found in the results.")
            else:
                st.warning("⚠️ No student found with that ID or Name.")

        if st.button("❌ Cancel Search"):
            st.session_state.delete_mode = False
            st.session_state.delete_candidate = None
            st.rerun()
